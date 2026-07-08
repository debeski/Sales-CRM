"""
Sales reporting — domain aggregates over real invoices (revenue, outstanding,
top products/services) plus an XLSX export.

This is distinct from DjangoLux's activity-log report overview: that reports *who
did what*; this reports *what was sold*. Money figures are in LYD (the frozen
per-invoice totals), so the report is stable regardless of later rate moves.
"""
from datetime import date, timedelta
from decimal import Decimal
from io import BytesIO

from django.db.models import Count, DecimalField, ExpressionWrapper, F, Sum
from django.db.models.functions import Coalesce, TruncMonth
from django.utils import timezone

from common.access import apply_ownership
from common.i18n import t
from finance.services import get_current_rate, usd_to_lyd

from .models import Invoice, InvoiceItem, Payment

LIVE_STATUSES = [Invoice.STATUS_ISSUED, Invoice.STATUS_PARTIAL, Invoice.STATUS_PAID]
_Z = Decimal("0.00")
_MONEY = DecimalField(max_digits=20, decimal_places=2)


def default_window():
    """(date_from, date_to) defaulting to the current month-to-date."""
    today = timezone.localdate()
    return today.replace(day=1), today


def parse_window(date_from, date_to):
    """Coerce two optional ISO date strings into a valid (from, to) pair."""
    from django.utils.dateparse import parse_date

    d1 = parse_date(date_from) if date_from else None
    d2 = parse_date(date_to) if date_to else None
    df, dt = default_window()
    d1 = d1 or df
    d2 = d2 or dt
    if d1 > d2:
        d1, d2 = d2, d1
    return d1, d2


def build_sales_report(date_from, date_to, actor=None):
    """Aggregate issued/partial/paid invoices within the window into a report dict.

    Row-scoped by ``actor``: a sales rep only ever reports on their own sales,
    while a manager holding ``view_all_invoice`` reports on the whole store
    (see ``common.access.apply_ownership``). ``actor=None`` is an unrestricted
    system/programmatic context (the web views always pass ``request.user``).
    """
    qs = Invoice.objects.all()
    if actor is not None:
        qs = apply_ownership(qs, actor)
    qs = qs.filter(status__in=LIVE_STATUSES, invoice_date__range=(date_from, date_to))

    agg = qs.aggregate(total=Sum("total_lyd"), paid=Sum("amount_paid"), count=Count("id"))
    total = agg["total"] or _Z
    paid = agg["paid"] or _Z

    by_status = list(
        qs.values("status").annotate(count=Count("id"), total=Sum("total_lyd")).order_by("status")
    )
    status_labels = dict(Invoice.STATUS_CHOICES)
    for row in by_status:
        row["status_key"] = f"status_{row['status']}"
        row["status_label"] = t(f"status_{row['status']}", status_labels.get(row["status"], row["status"]))

    daily = list(
        qs.values("invoice_date").annotate(count=Count("id"), total=Sum("total_lyd")).order_by("invoice_date")
    )

    items = InvoiceItem.objects.filter(invoice__in=qs)
    by_product = list(
        items.filter(kind=InvoiceItem.KIND_PRODUCT)
        .values("description")
        .annotate(qty=Sum("quantity"), total=Sum("line_total_lyd"))
        .order_by("-total")[:50]
    )
    by_service = list(
        items.filter(kind=InvoiceItem.KIND_SERVICE)
        .values("description")
        .annotate(qty=Sum("quantity"), total=Sum("line_total_lyd"))
        .order_by("-total")[:50]
    )

    invoices = list(
        qs.order_by("invoice_date", "number").values(
            "number", "invoice_date", "customer_name", "status", "total_lyd", "amount_paid"
        )
    )
    for inv in invoices:
        inv["balance"] = (inv["total_lyd"] or _Z) - (inv["amount_paid"] or _Z)
        inv["status_label"] = t(f"status_{inv['status']}", status_labels.get(inv["status"], inv["status"]))

    return {
        "date_from": date_from,
        "date_to": date_to,
        "total_sales": total,
        "total_paid": paid,
        "outstanding": total - paid,
        "invoice_count": agg["count"] or 0,
        "by_status": by_status,
        "daily": daily,
        "by_product": by_product,
        "by_service": by_service,
        "invoices": invoices,
    }


def build_sales_report_xlsx(report):
    """Render a report dict to an .xlsx workbook and return raw bytes."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0D6EFD")
    title_font = Font(bold=True, size=14)
    money_fmt = "#,##0.00"

    def style_header(ws, row=1):
        for cell in ws[row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

    def autosize(ws):
        for col in ws.columns:
            width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(width + 3, 50)

    # --- Summary sheet ---
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Switch — Sales Report"
    ws["A1"].font = title_font
    ws["A2"] = f"Period: {report['date_from']} → {report['date_to']}"
    rows = [
        ("Total Sales (LYD)", report["total_sales"]),
        ("Total Collected (LYD)", report["total_paid"]),
        ("Outstanding (LYD)", report["outstanding"]),
        ("Invoices", report["invoice_count"]),
    ]
    for i, (label, value) in enumerate(rows, start=4):
        ws[f"A{i}"] = label
        ws[f"A{i}"].font = Font(bold=True)
        ws[f"B{i}"] = value
        if "LYD" in label:
            ws[f"B{i}"].number_format = money_fmt
    autosize(ws)

    def add_sheet(title, headers, data_rows, money_cols=()):
        sheet = wb.create_sheet(title)
        sheet.append(headers)
        style_header(sheet)
        for r in data_rows:
            sheet.append(r)
        for col_idx in money_cols:
            for row in sheet.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    cell.number_format = money_fmt
        autosize(sheet)
        return sheet

    add_sheet(
        "By Status",
        ["Status", "Count", "Total (LYD)"],
        [(r["status_label"], r["count"], r["total"]) for r in report["by_status"]],
        money_cols=(3,),
    )
    add_sheet(
        "Daily",
        ["Date", "Invoices", "Total (LYD)"],
        [(str(r["invoice_date"]), r["count"], r["total"]) for r in report["daily"]],
        money_cols=(3,),
    )
    add_sheet(
        "Top Products",
        ["Product", "Qty", "Total (LYD)"],
        [(r["description"], r["qty"], r["total"]) for r in report["by_product"]],
        money_cols=(3,),
    )
    add_sheet(
        "Top Services",
        ["Service", "Qty", "Total (LYD)"],
        [(r["description"], r["qty"], r["total"]) for r in report["by_service"]],
        money_cols=(3,),
    )
    add_sheet(
        "Invoices",
        ["Number", "Date", "Customer", "Status", "Total (LYD)", "Paid (LYD)", "Balance (LYD)"],
        [
            (
                r["number"], str(r["invoice_date"]), r["customer_name"], r["status_label"],
                r["total_lyd"], r["amount_paid"], r["balance"],
            )
            for r in report["invoices"]
        ],
        money_cols=(5, 6, 7),
    )

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# --------------------------------------------------------------------------- #
# Fiscal-year financial report (whole-store P&L — never row-scoped)
# --------------------------------------------------------------------------- #
def fiscal_year_window(year):
    """A fiscal year in Libya is the calendar year: Jan 1 – Dec 31 of ``year``."""
    return date(year, 1, 1), date(year, 12, 31)


def available_fiscal_years():
    """Years to offer in the picker: from the first invoice year to this year."""
    this_year = timezone.localdate().year
    first = Invoice.objects.order_by("invoice_date").values_list("invoice_date", flat=True).first()
    start = first.year if first else this_year
    return list(range(this_year, start - 1, -1))


def build_financial_report(date_from, date_to):
    """Whole-store financial summary for a period (a fiscal year by default).

    Deliberately NOT ``apply_ownership``-scoped — this is an owner/manager P&L
    over the entire business, gated by ``sales.view_financial_report``.

    Period figures: revenue, COGS (estimate), gross profit, cash collected.
    Snapshot (current) figures: outstanding receivables, inventory value — these
    are point-in-time, labelled as such in the template.

    COGS uses the unit cost **frozen on each invoice line at the time of sale**
    (``InvoiceItem.unit_cost_usd``), converted at that invoice's own frozen rate —
    so it's exact. Lines created before cost-freezing fall back to the product's
    current cost via ``Coalesce``.
    """
    from catalog.models import Product

    live = Invoice.objects.filter(status__in=LIVE_STATUSES, invoice_date__range=(date_from, date_to))
    agg = live.aggregate(revenue=Sum("total_lyd"), count=Count("id"))
    revenue = agg["revenue"] or _Z

    # COGS = qty × frozen unit cost (fallback: current product cost) × frozen rate.
    cogs_expr = ExpressionWrapper(
        F("quantity")
        * Coalesce(F("unit_cost_usd"), F("product__cost_usd"))
        * F("invoice__exchange_rate"),
        output_field=_MONEY,
    )
    cogs = (
        InvoiceItem.objects.filter(
            invoice__in=live, kind=InvoiceItem.KIND_PRODUCT, product__isnull=False
        ).aggregate(c=Sum(cogs_expr))["c"]
    ) or _Z
    gross_profit = revenue - cogs
    margin = (gross_profit / revenue * Decimal("100")) if revenue else _Z

    # Cash actually collected in the period.
    cash_collected = (
        Payment.objects.filter(paid_at__date__range=(date_from, date_to)).aggregate(s=Sum("amount"))["s"]
    ) or _Z

    # Current outstanding receivables (issued/partial, any date).
    open_qs = Invoice.objects.filter(status__in=[Invoice.STATUS_ISSUED, Invoice.STATUS_PARTIAL])
    open_agg = open_qs.aggregate(total=Sum("total_lyd"), paid=Sum("amount_paid"))
    receivables = (open_agg["total"] or _Z) - (open_agg["paid"] or _Z)

    # Current closing-stock value (Σ stock × unit cost), USD → LYD at live rate.
    rate = get_current_rate()
    inv_value_usd = (
        Product.objects.filter(is_active=True, track_stock=True).aggregate(
            v=Sum(ExpressionWrapper(F("stock_qty") * F("cost_usd"), output_field=_MONEY))
        )["v"]
    ) or _Z
    inventory_value = usd_to_lyd(inv_value_usd, rate)

    # Revenue by month for a trend table.
    monthly = [
        {"month": row["m"], "total": row["t"] or _Z}
        for row in live.annotate(m=TruncMonth("invoice_date")).values("m").annotate(t=Sum("total_lyd")).order_by("m")
    ]

    return {
        "date_from": date_from,
        "date_to": date_to,
        "year": date_from.year if date_from.year == date_to.year else None,
        "revenue": revenue,
        "invoice_count": agg["count"] or 0,
        "cogs": cogs,
        "gross_profit": gross_profit,
        "margin_percent": margin,
        "cash_collected": cash_collected,
        "receivables": receivables,
        "inventory_value": inventory_value,
        "monthly": monthly,
    }
